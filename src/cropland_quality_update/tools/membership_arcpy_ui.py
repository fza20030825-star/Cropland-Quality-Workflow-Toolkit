"""ArcPy UI tool for calculating cropland quality memberships and grades."""

from __future__ import annotations

import json
import logging
import math
import queue
import shutil
import threading
import traceback
import uuid
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import BOTH, END, LEFT, IntVar, StringVar, Text, Tk, Toplevel, filedialog, messagebox, ttk

try:
    import arcpy
except ImportError as exc:  # pragma: no cover - shown to user at runtime
    arcpy = None
    ARCPY_IMPORT_ERROR = exc
else:
    ARCPY_IMPORT_ERROR = None

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - shown to user at runtime
    load_workbook = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None

from cropland_quality_update.paths import resolve_paths
from cropland_quality_update.tools import vector_common_arcpy as vector_tool


DEFAULT_AREA_NAME = "秦岭大巴山林农区"
ALTITUDE_REQUIRED_AREAS = {"秦岭大巴山林农区", "渝鄂湘黔边境山地林农牧区"}
ISSUE_DETAIL_LIMIT = 50
SAMPLE_FIELD_VALUE_LIMIT = 10
GEOMETRY_FIELD_NAMES = {"shape", "shape_length", "shape_area", "shape_leng"}
SKIP_FIELD_TYPES = {"OID", "Geometry", "Blob", "Raster", "GUID", "GlobalID", "XML"}
TEXT_FIELD_TYPES = {"String"}
NUMERIC_FIELD_TYPES = {"SmallInteger", "Integer", "Single", "Double"}
RESULT_SCORE_FIELD = "评价得分"
RESULT_GRADE_FIELD = "质量等级"
OVERLAP_MARK_FIELD = "是否重叠"
OVERLAP_MODE_MARK = "mark"
OVERLAP_MODE_ORDER_ERASE = "order_erase"
TEMP_CANDIDATE_ID_FIELD = "HS_ID"
TEMP_SOURCE_ORDER_FIELD = "HS_SRC"
TEMP_SOURCE_OID_FIELD = "HS_OID"
STANDARD_INDICATOR_FIELDS = [
    "地形部位",
    "耕层质地",
    "水资源条件",
    "排水能力",
    "海拔高度",
    "有机质",
    "有效土层厚度",
    "土壤容重",
    "速效钾",
    "有效磷",
    "质地构型",
    "酸碱度",
    "耕层厚度",
]
INDICATOR_TEXT_FIELDS = {"地形部位", "耕层质地", "水资源条件", "排水能力", "质地构型"}


@dataclass(frozen=True)
class VectorSource:
    kind: str  # shp | gdb
    source_path: Path
    layer_name: str | None
    display_name: str


@dataclass(frozen=True)
class NumericRule:
    indicator: str
    function_code: str
    function_type: str
    a: float
    b: float | None
    c: float | None
    lower: float
    upper: float


@dataclass(frozen=True)
class GradeRule:
    grade_value: int
    grade_name: str
    lower: float | None
    upper: float | None


@dataclass(frozen=True)
class RuleSet:
    area_name: str
    rule_path: Path
    weights: dict[str, float]
    concept_memberships: dict[str, dict[str, float]]
    numeric_rules: dict[str, NumericRule]
    grade_rules: list[GradeRule]

    @property
    def indicators(self) -> list[str]:
        return list(self.weights.keys())

    @property
    def concept_indicators(self) -> set[str]:
        return set(self.concept_memberships.keys())

    @property
    def numeric_indicators(self) -> set[str]:
        return set(self.numeric_rules.keys())


@dataclass(frozen=True)
class FieldBinding:
    indicator: str
    source_field: str
    field_type: str


@dataclass(frozen=True)
class FeatureSample:
    oid: int
    field_values: list[tuple[str, str]]


@dataclass(frozen=True)
class MissingValueIssue:
    indicator: str
    field_name: str
    count: int
    samples: list[FeatureSample]


@dataclass(frozen=True)
class InvalidCategoryIssue:
    indicator: str
    field_name: str
    invalid_value: str
    count: int
    samples: list[FeatureSample]


@dataclass(frozen=True)
class ValidationReport:
    ok: bool
    source: VectorSource
    rule_set: RuleSet
    feature_count: int
    field_bindings: dict[str, FieldBinding]
    missing_fields: list[str]
    type_errors: list[str]
    missing_value_issues: list[MissingValueIssue]
    invalid_category_issues: list[InvalidCategoryIssue]
    text: str


@dataclass(frozen=True)
class SourceValidationResult:
    source: VectorSource
    feature_count: int
    field_bindings: dict[str, FieldBinding]
    missing_fields: list[str]
    ambiguous_fields: list[str]
    type_errors: list[str]
    missing_value_issues: list[MissingValueIssue]
    invalid_category_issues: list[InvalidCategoryIssue]


@dataclass(frozen=True)
class HighStandardPreflightReport:
    ok: bool
    input_sources: list[VectorSource]
    rule_set: RuleSet
    projection_infos: list[object]
    projections_same: bool
    target_projection_source: VectorSource
    target_spatial_reference: object
    source_results: list[SourceValidationResult]
    feature_count: int
    text: str


@dataclass(frozen=True)
class CalculationJob:
    job_id: str
    source: VectorSource
    output_path: Path
    output_feature_name: str | None
    output_kind: str
    rule_set: RuleSet
    field_bindings: dict[str, FieldBinding]
    created_at: str
    validation_report: str


@dataclass(frozen=True)
class HighStandardCalculationJob:
    job_id: str
    input_sources: list[VectorSource]
    output_path: Path
    output_feature_name: str | None
    output_kind: str
    overlap_mode: str
    rule_set: RuleSet
    source_results: list[SourceValidationResult]
    target_spatial_reference: object
    target_projection_source: VectorSource
    created_at: str
    validation_report: str


@dataclass(frozen=True)
class HighStandardFeatureCandidate:
    candidate_id: int
    source_label: str
    source_order: int
    source_oid: int
    geometry: object
    raw_values: list[object]
    membership_values: list[float | None]
    score: float
    grade: int


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def timestamp_for_file() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def require_runtime() -> None:
    missing = []
    if arcpy is None:
        missing.append(f"arcpy ({ARCPY_IMPORT_ERROR})")
    if load_workbook is None:
        missing.append(f"openpyxl ({OPENPYXL_IMPORT_ERROR})")
    if missing:
        raise RuntimeError("缺少运行包：" + "；".join(missing))


def make_vector_source(kind: str, source_path: Path, layer_name: str | None = None) -> VectorSource:
    source_path = source_path.resolve()
    if kind == "shp":
        shp_path = source_path if source_path.suffix.lower() == ".shp" else source_path.with_suffix(".shp")
        return VectorSource("shp", shp_path, None, str(shp_path))
    if kind == "gdb":
        if not layer_name:
            raise ValueError("gdb source requires layer_name")
        return VectorSource("gdb", source_path, layer_name, f"{source_path}\\{layer_name}")
    raise ValueError(f"unsupported source kind: {kind}")


def source_label(source: VectorSource) -> str:
    return source.display_name


def source_dataset_path(source: VectorSource) -> str:
    if source.kind == "gdb":
        return str(source.source_path / str(source.layer_name))
    return str(source.source_path)


def is_gdb_path(path: Path) -> bool:
    return path.suffix.lower() == ".gdb" and path.is_dir()


def find_nearest_gdb_path(path: Path) -> Path | None:
    current = path.resolve()
    for candidate in (current, *current.parents):
        if candidate.suffix.lower() == ".gdb" and candidate.is_dir():
            return candidate
    return None


def list_gdb_polygon_layers(gdb_path: Path) -> list[VectorSource]:
    require_runtime()
    previous_workspace = arcpy.env.workspace
    arcpy.env.workspace = str(gdb_path)
    try:
        sources: list[VectorSource] = []
        for feature_class in arcpy.ListFeatureClasses(feature_type="Polygon") or []:
            sources.append(make_vector_source("gdb", gdb_path, feature_class))
        for dataset in arcpy.ListDatasets(feature_type="Feature") or []:
            for feature_class in arcpy.ListFeatureClasses(feature_dataset=dataset, feature_type="Polygon") or []:
                sources.append(make_vector_source("gdb", gdb_path, f"{dataset}\\{feature_class}"))
        return sources
    finally:
        arcpy.env.workspace = previous_workspace


def output_format(path: Path) -> str:
    return "shp" if path.suffix.lower() == ".shp" else "gdb"


def output_dataset_path(output_kind: str, output_path: Path, output_feature_name: str | None) -> Path:
    return output_path if output_kind == "shp" else output_path / str(output_feature_name)


def validate_gdb_feature_name(feature_name: str, gdb_path: Path) -> str:
    workspace = str(gdb_path if gdb_path.exists() else gdb_path.parent)
    candidate = arcpy.ValidateTableName(feature_name, workspace)
    if not candidate:
        raise RuntimeError("输出面要素类名称无效。")
    if not candidate[0].isalpha():
        candidate = arcpy.ValidateTableName(f"Step_{candidate}", workspace)
    return candidate


def sidecar_paths(shp_path: Path) -> list[Path]:
    base = shp_path.with_suffix("")
    return [base.with_suffix(ext) for ext in (".shp", ".shx", ".dbf", ".prj", ".cpg", ".shp.xml", ".fields.json")]


def delete_output_dataset(output_path: Path, output_feature_name: str | None = None) -> None:
    require_runtime()
    if output_format(output_path) == "shp":
        for path in sidecar_paths(output_path):
            if path.exists():
                path.unlink()
        return
    if output_feature_name:
        feature_class = str(output_path / output_feature_name)
        if arcpy.Exists(feature_class):
            arcpy.management.Delete(feature_class)
        return
    if output_path.exists():
        shutil.rmtree(output_path)


def is_data_field(field) -> bool:
    if field.type in SKIP_FIELD_TYPES:
        return False
    if field.name.lower() in GEOMETRY_FIELD_NAMES:
        return False
    return True


def field_value_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    if len(text) > 80:
        return text[:77] + "..."
    return text


def is_blank_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def normalize_category(value) -> str:
    return str(value).strip()


def oid_field_name(dataset_path: str) -> str:
    return arcpy.Describe(dataset_path).OIDFieldName


def first_data_field_names(dataset_path: str, limit: int = SAMPLE_FIELD_VALUE_LIMIT) -> list[str]:
    names: list[str] = []
    for field in arcpy.ListFields(dataset_path):
        if is_data_field(field):
            names.append(field.name)
        if len(names) >= limit:
            break
    return names


def data_field_names(dataset_path: str) -> list[str]:
    return [field.name for field in arcpy.ListFields(dataset_path) if is_data_field(field)]


def fetch_field_values_by_oid(dataset_path: str, oid_value: int, field_names: list[str]) -> list[tuple[str, str]]:
    if oid_value is None or oid_value < 0 or not field_names:
        return []
    oid_name = oid_field_name(dataset_path)
    where = f"{arcpy.AddFieldDelimiters(dataset_path, oid_name)} = {int(oid_value)}"
    with arcpy.da.SearchCursor(dataset_path, field_names, where_clause=where) as cursor:
        for row in cursor:
            return [(name, field_value_text(value)) for name, value in zip(field_names, row)]
    return []


def make_issue_samples(dataset_path: str, oids: list[int]) -> list[FeatureSample]:
    sample_fields = first_data_field_names(dataset_path)
    return [FeatureSample(oid, fetch_field_values_by_oid(dataset_path, oid, sample_fields)) for oid in oids]


def read_shp_field_mapping(source: VectorSource) -> dict[str, str]:
    if source.kind != "shp":
        return {}
    mapping_path = source.source_path.with_suffix(".fields.json")
    if not mapping_path.exists():
        return {}
    try:
        raw = json.loads(mapping_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {str(original): str(actual) for original, actual in raw.items()}


def normalize_indicator_key(value: str) -> str:
    text = str(value or "").strip()
    for token in (" ", "\t", "\r", "\n", "_", "-", "（", "）", "(", ")", "　"):
        text = text.replace(token, "")
    if text.lower() == "ph":
        return "酸碱度"
    return text.lower()


def possible_indicator_matches(label: str, indicators: list[str]) -> set[str]:
    key = normalize_indicator_key(label)
    if not key:
        return set()
    if key == "酸碱度":
        return {"酸碱度"} if "酸碱度" in indicators else set()
    matches: set[str] = set()
    for indicator in indicators:
        indicator_key = normalize_indicator_key(indicator)
        if key == indicator_key or indicator_key.startswith(key) or key.startswith(indicator_key):
            matches.add(indicator)
    return matches


def field_labels_by_actual(source: VectorSource, fields: dict[str, object]) -> dict[str, set[str]]:
    labels: dict[str, set[str]] = {name: {name} for name in fields}
    for name, field in fields.items():
        alias = getattr(field, "aliasName", "") or ""
        if alias and alias != name:
            labels[name].add(alias)
    for original, actual in read_shp_field_mapping(source).items():
        if actual in labels:
            labels[actual].add(original)
    return labels


def bind_indicator_fields_loose(
    source: VectorSource,
    indicators: list[str],
    optional_indicators: set[str],
) -> tuple[dict[str, FieldBinding], list[str], list[str]]:
    dataset = source_dataset_path(source)
    fields = {field.name: field for field in arcpy.ListFields(dataset) if is_data_field(field)}
    labels = field_labels_by_actual(source, fields)
    candidates: dict[str, set[str]] = {indicator: set() for indicator in indicators}
    ambiguous_labels: list[str] = []
    for actual_name, names in labels.items():
        matched_by_label: set[str] = set()
        for label in names:
            matched_by_label |= possible_indicator_matches(label, indicators)
        if len(matched_by_label) == 1:
            indicator = next(iter(matched_by_label))
            candidates[indicator].add(actual_name)
        elif len(matched_by_label) > 1:
            label_text = "、".join(sorted(names))
            ambiguous_labels.append(f"{actual_name}({label_text}) -> {'、'.join(sorted(matched_by_label))}")

    bindings: dict[str, FieldBinding] = {}
    missing: list[str] = []
    ambiguous: list[str] = []
    for indicator in indicators:
        actual_candidates = sorted(candidates[indicator])
        if len(actual_candidates) == 1:
            field = fields[actual_candidates[0]]
            bindings[indicator] = FieldBinding(indicator, actual_candidates[0], field.type)
        elif len(actual_candidates) > 1:
            ambiguous.append(f"{indicator}: {'、'.join(actual_candidates)}")
        elif indicator not in optional_indicators:
            missing.append(indicator)
    ambiguous.extend(ambiguous_labels)
    return bindings, missing, ambiguous


def find_rule_file(rules_dir: Path, area_name: str) -> Path:
    candidates = [
        path
        for path in sorted(rules_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and path.stem.split("_", 1)[0] == area_name
    ]
    if not candidates:
        raise RuntimeError(f"没有找到国标二级农业区对应规则文件：{area_name}_*.xlsx")
    if len(candidates) > 1:
        names = "；".join(path.name for path in candidates)
        raise RuntimeError(f"找到多个规则文件，请保留一个：{names}")
    return candidates[0]


def list_rule_area_options(rules_dir: Path) -> list[str]:
    area_names: list[str] = []
    seen: set[str] = set()
    for path in sorted(rules_dir.glob("*.xlsx")):
        if path.name.startswith("~$") or "_" not in path.stem:
            continue
        area_name = path.stem.split("_", 1)[0].strip()
        if not area_name or area_name in seen:
            continue
        seen.add(area_name)
        area_names.append(area_name)
    return area_names


def sheet_rows(ws) -> list[dict[str, object]]:
    headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, object]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() != "" for value in raw):
            continue
        rows.append({header: raw[index] if index < len(raw) else None for index, header in enumerate(headers) if header})
    return rows


def as_float(value, field_name: str) -> float:
    if value is None or str(value).strip() == "":
        raise RuntimeError(f"规则表字段为空：{field_name}")
    return float(value)


def as_optional_float(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    return float(value)


def load_rule_set(rules_dir: Path, area_name: str) -> RuleSet:
    require_runtime()
    rule_path = find_rule_file(rules_dir, area_name)
    wb = load_workbook(rule_path, data_only=True, read_only=True)
    for sheet_name in ("指标权重", "概念指标隶属度", "数值指标隶属函数", "等级划分指数"):
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"规则文件缺少 sheet：{sheet_name}")

    weights: dict[str, float] = {}
    for row in sheet_rows(wb["指标权重"]):
        indicator = str(row.get("指标名称", "")).strip()
        if indicator:
            weights[indicator] = as_float(row.get("指标权重"), f"{indicator}.指标权重")
    if not weights:
        raise RuntimeError("规则表“指标权重”中没有读取到任何指标。")

    concept_memberships: dict[str, dict[str, float]] = {}
    for row in sheet_rows(wb["概念指标隶属度"]):
        indicator = str(row.get("指标名称", "")).strip()
        category = normalize_category(row.get("类别值", ""))
        if indicator and category:
            concept_memberships.setdefault(indicator, {})[category] = as_float(row.get("隶属度"), f"{indicator}.{category}.隶属度")

    numeric_rules: dict[str, NumericRule] = {}
    for row in sheet_rows(wb["数值指标隶属函数"]):
        indicator = str(row.get("指标名称", "")).strip()
        if not indicator:
            continue
        numeric_rules[indicator] = NumericRule(
            indicator=indicator,
            function_code=str(row.get("函数代码", "")).strip(),
            function_type=str(row.get("函数类型", "")).strip(),
            a=as_float(row.get("a"), f"{indicator}.a"),
            b=as_optional_float(row.get("b")),
            c=as_optional_float(row.get("c")),
            lower=as_float(row.get("下限"), f"{indicator}.下限"),
            upper=as_float(row.get("上限"), f"{indicator}.上限"),
        )

    grade_rules: list[GradeRule] = []
    for row in sheet_rows(wb["等级划分指数"]):
        grade_value = row.get("等级值")
        if grade_value is None:
            continue
        grade_rules.append(
            GradeRule(
                grade_value=int(grade_value),
                grade_name=str(row.get("等级名称", "")).strip(),
                lower=as_optional_float(row.get("下限")),
                upper=as_optional_float(row.get("上限")),
            )
        )
    grade_rules.sort(key=lambda item: item.grade_value)

    missing_rules = set(weights) - set(concept_memberships) - set(numeric_rules)
    extra_rules = (set(concept_memberships) | set(numeric_rules)) - set(weights)
    if missing_rules:
        raise RuntimeError("规则表中这些指标没有隶属度规则：" + "、".join(sorted(missing_rules)))
    if extra_rules:
        raise RuntimeError("规则表中这些隶属度规则没有对应权重：" + "、".join(sorted(extra_rules)))
    return RuleSet(area_name, rule_path, weights, concept_memberships, numeric_rules, grade_rules)


def bind_indicator_fields(source: VectorSource, rule_set: RuleSet) -> tuple[dict[str, FieldBinding], list[str]]:
    dataset = source_dataset_path(source)
    fields = {field.name: field for field in arcpy.ListFields(dataset) if is_data_field(field)}
    shp_mapping = read_shp_field_mapping(source)
    bindings: dict[str, FieldBinding] = {}
    missing: list[str] = []
    for indicator in rule_set.indicators:
        source_field = indicator if indicator in fields else shp_mapping.get(indicator, "")
        if source_field and source_field in fields:
            bindings[indicator] = FieldBinding(indicator, source_field, fields[source_field].type)
        else:
            missing.append(indicator)
    return bindings, missing


def collect_validation_issues(
    source: VectorSource,
    rule_set: RuleSet,
    bindings: dict[str, FieldBinding],
) -> tuple[list[str], list[MissingValueIssue], list[InvalidCategoryIssue]]:
    dataset = source_dataset_path(source)
    oid_name = oid_field_name(dataset)
    type_errors: list[str] = []
    missing_oids: dict[str, list[int]] = {indicator: [] for indicator in bindings}
    invalid_categories: dict[tuple[str, str], list[int]] = {}

    for indicator in rule_set.concept_indicators:
        binding = bindings.get(indicator)
        if binding and binding.field_type not in TEXT_FIELD_TYPES:
            type_errors.append(f"类别指标字段必须为文本型：{indicator} -> {binding.source_field}（实际 {binding.field_type}）")
    for indicator in rule_set.numeric_indicators:
        binding = bindings.get(indicator)
        if binding and binding.field_type not in NUMERIC_FIELD_TYPES:
            type_errors.append(f"数值指标字段必须为数值型：{indicator} -> {binding.source_field}（实际 {binding.field_type}）")

    cursor_fields = [oid_name] + [binding.source_field for binding in bindings.values()]
    binding_by_cursor_index = list(bindings.values())
    allowed_categories = {
        indicator: set(values.keys()) for indicator, values in rule_set.concept_memberships.items()
    }
    with arcpy.da.SearchCursor(dataset, cursor_fields) as cursor:
        for row in cursor:
            oid = int(row[0])
            for value, binding in zip(row[1:], binding_by_cursor_index):
                if is_blank_value(value):
                    missing_oids[binding.indicator].append(oid)
                    continue
                if binding.indicator in allowed_categories:
                    text = normalize_category(value)
                    if text not in allowed_categories[binding.indicator]:
                        invalid_categories.setdefault((binding.indicator, text), []).append(oid)

    missing_total = sum(len(oids) for oids in missing_oids.values())
    invalid_total = sum(len(oids) for oids in invalid_categories.values())
    show_missing_details = missing_total < ISSUE_DETAIL_LIMIT
    show_invalid_details = invalid_total < ISSUE_DETAIL_LIMIT

    missing_issues: list[MissingValueIssue] = []
    for indicator, oids in missing_oids.items():
        if not oids:
            continue
        binding = bindings[indicator]
        samples = make_issue_samples(dataset, oids) if show_missing_details else []
        missing_issues.append(MissingValueIssue(indicator, binding.source_field, len(oids), samples))

    invalid_issues: list[InvalidCategoryIssue] = []
    for (indicator, invalid_value), oids in sorted(invalid_categories.items()):
        binding = bindings[indicator]
        samples = make_issue_samples(dataset, oids) if show_invalid_details else []
        invalid_issues.append(InvalidCategoryIssue(indicator, binding.source_field, invalid_value, len(oids), samples))

    return type_errors, missing_issues, invalid_issues


def standard_indicator_type_errors(bindings: dict[str, FieldBinding]) -> list[str]:
    errors: list[str] = []
    for indicator, binding in bindings.items():
        if indicator in INDICATOR_TEXT_FIELDS:
            if binding.field_type not in TEXT_FIELD_TYPES:
                errors.append(f"概念指标字段必须为文本型：{indicator} -> {binding.source_field}（实际 {binding.field_type}）")
        elif binding.field_type not in NUMERIC_FIELD_TYPES:
            errors.append(f"数值指标字段必须为数值型：{indicator} -> {binding.source_field}（实际 {binding.field_type}）")
    return errors


def format_samples(samples: list[FeatureSample]) -> list[str]:
    lines: list[str] = []
    for index, sample in enumerate(samples, start=1):
        values = "；".join(f"{name}={value}" for name, value in sample.field_values) or "无可展示字段值"
        lines.append(f"       {index}. OID={sample.oid}；{values}")
    return lines


def build_validation_text(report: ValidationReport) -> str:
    lines: list[str] = []
    lines.append("隶属度计算前检查报告")
    lines.append("=" * 60)
    lines.append(f"输入数据：{source_label(report.source)}")
    lines.append(f"国标二级农业区：{report.rule_set.area_name}")
    lines.append(f"规则文件：{report.rule_set.rule_path}")
    lines.append(f"面要素数：{report.feature_count}")
    lines.append(f"规则指标数：{len(report.rule_set.weights)}")
    lines.append("")
    lines.append("一、字段匹配")
    if report.missing_fields:
        lines.append(f"缺少以下规则要求的 {len(report.rule_set.weights)} 个指标字段，已打回：")
        for name in report.missing_fields:
            lines.append(f"   - {name}")
    else:
        lines.append(f"规则要求的 {len(report.rule_set.weights)} 个指标字段均已找到。")
        for indicator in report.rule_set.indicators:
            binding = report.field_bindings[indicator]
            lines.append(f"   - {indicator} -> {binding.source_field}（{binding.field_type}）")
    lines.append("")
    lines.append("二、字段类型检查")
    if report.type_errors:
        lines.append("发现字段类型不符合规则，已打回：")
        lines.extend(f"   - {item}" for item in report.type_errors)
    else:
        lines.append("字段类型检查通过。")
    lines.append("")
    lines.append("三、空值检查")
    missing_total = sum(issue.count for issue in report.missing_value_issues)
    if missing_total:
        lines.append(f"发现空值 {missing_total} 个，已打回。")
        if missing_total >= ISSUE_DETAIL_LIMIT:
            lines.append(f"空值过多（总体空值 {missing_total} 个），不展开字段值。")
        for issue in report.missing_value_issues:
            lines.append(f"   - 指标：{issue.indicator}；字段：{issue.field_name}；空值要素数：{issue.count}")
            lines.extend(format_samples(issue.samples))
    else:
        lines.append("未发现规则指标字段空值。")
    lines.append("")
    lines.append("四、类别值检查")
    invalid_total = sum(issue.count for issue in report.invalid_category_issues)
    if invalid_total:
        lines.append(f"发现类别值不在规则表中 {invalid_total} 个，已打回。")
        if invalid_total >= ISSUE_DETAIL_LIMIT:
            lines.append(f"类别值错误过多（总体错误 {invalid_total} 个），不展开字段值。")
        for issue in report.invalid_category_issues:
            lines.append(
                f"   - 指标：{issue.indicator}；字段：{issue.field_name}；非法类别值：{issue.invalid_value}；要素数：{issue.count}"
            )
            allowed = "、".join(sorted(report.rule_set.concept_memberships[issue.indicator]))
            lines.append(f"     允许值：{allowed}")
            lines.extend(format_samples(issue.samples))
    else:
        lines.append("类别字段取值均能在规则表中匹配。")
    lines.append("")
    if report.ok:
        lines.append("检查通过，可以继续计算隶属度、评价得分和质量等级。")
    else:
        lines.append("检查未通过，请修正数据后重新检查。")
    return "\n".join(lines)


def validate_source(source: VectorSource, rule_set: RuleSet) -> ValidationReport:
    require_runtime()
    dataset = source_dataset_path(source)
    if not arcpy.Exists(dataset):
        raise RuntimeError(f"输入数据不存在：{dataset}")
    desc = arcpy.Describe(dataset)
    if getattr(desc, "shapeType", "") != "Polygon":
        raise RuntimeError("输入数据必须是面矢量。")
    feature_count = int(arcpy.management.GetCount(dataset)[0])
    bindings, missing_fields = bind_indicator_fields(source, rule_set)
    type_errors: list[str] = []
    missing_issues: list[MissingValueIssue] = []
    invalid_issues: list[InvalidCategoryIssue] = []
    if not missing_fields:
        type_errors, missing_issues, invalid_issues = collect_validation_issues(source, rule_set, bindings)
    ok = not missing_fields and not type_errors and not missing_issues and not invalid_issues
    draft = ValidationReport(
        ok=ok,
        source=source,
        rule_set=rule_set,
        feature_count=feature_count,
        field_bindings=bindings,
        missing_fields=missing_fields,
        type_errors=type_errors,
        missing_value_issues=missing_issues,
        invalid_category_issues=invalid_issues,
        text="",
    )
    return ValidationReport(
        ok=draft.ok,
        source=draft.source,
        rule_set=draft.rule_set,
        feature_count=draft.feature_count,
        field_bindings=draft.field_bindings,
        missing_fields=draft.missing_fields,
        type_errors=draft.type_errors,
        missing_value_issues=draft.missing_value_issues,
        invalid_category_issues=draft.invalid_category_issues,
        text=build_validation_text(draft),
    )


def optional_indicators_for_area(area_name: str) -> set[str]:
    return set() if area_name in ALTITUDE_REQUIRED_AREAS else {"海拔高度"}


def validate_high_standard_sources(
    sources: list[VectorSource],
    rule_set: RuleSet,
    target_projection_source: VectorSource,
    target_spatial_reference: object,
) -> HighStandardPreflightReport:
    require_runtime()
    optional_indicators = optional_indicators_for_area(rule_set.area_name)
    projection_infos, projections_same = vector_tool.analyze_projection_state(sources)
    source_results: list[SourceValidationResult] = []
    total_features = 0
    for source in sources:
        dataset = source_dataset_path(source)
        if not arcpy.Exists(dataset):
            raise RuntimeError(f"输入数据不存在：{source_label(source)}")
        desc = arcpy.Describe(dataset)
        if getattr(desc, "shapeType", "") != "Polygon":
            raise RuntimeError(f"输入数据必须是面矢量：{source_label(source)}")
        feature_count = int(arcpy.management.GetCount(dataset)[0])
        total_features += feature_count
        bindings, missing_fields, ambiguous_fields = bind_indicator_fields_loose(source, STANDARD_INDICATOR_FIELDS, optional_indicators)
        type_errors: list[str] = []
        missing_issues: list[MissingValueIssue] = []
        invalid_issues: list[InvalidCategoryIssue] = []
        if not missing_fields and not ambiguous_fields:
            type_errors = standard_indicator_type_errors(bindings)
            if not type_errors:
                type_errors, missing_issues, invalid_issues = collect_validation_issues(source, rule_set, bindings)
        source_results.append(
            SourceValidationResult(
                source=source,
                feature_count=feature_count,
                field_bindings=bindings,
                missing_fields=missing_fields,
                ambiguous_fields=ambiguous_fields,
                type_errors=type_errors,
                missing_value_issues=missing_issues,
                invalid_category_issues=invalid_issues,
            )
        )
    ok = all(
        not result.missing_fields
        and not result.ambiguous_fields
        and not result.type_errors
        and not result.missing_value_issues
        and not result.invalid_category_issues
        for result in source_results
    ) and all(info.spatial_reference is not None for info in projection_infos)
    draft = HighStandardPreflightReport(
        ok=ok,
        input_sources=sources,
        rule_set=rule_set,
        projection_infos=projection_infos,
        projections_same=projections_same,
        target_projection_source=target_projection_source,
        target_spatial_reference=target_spatial_reference,
        source_results=source_results,
        feature_count=total_features,
        text="",
    )
    return HighStandardPreflightReport(
        ok=draft.ok,
        input_sources=draft.input_sources,
        rule_set=draft.rule_set,
        projection_infos=draft.projection_infos,
        projections_same=draft.projections_same,
        target_projection_source=draft.target_projection_source,
        target_spatial_reference=draft.target_spatial_reference,
        source_results=draft.source_results,
        feature_count=draft.feature_count,
        text=build_high_standard_preflight_text(draft),
    )


def format_issue_block(result: SourceValidationResult, rule_set: RuleSet) -> list[str]:
    lines: list[str] = []
    if result.missing_fields:
        lines.append("   缺少字段：")
        lines.extend(f"      - {name}" for name in result.missing_fields)
    if result.ambiguous_fields:
        lines.append("   字段匹配歧义：")
        lines.extend(f"      - {item}" for item in result.ambiguous_fields[:ISSUE_DETAIL_LIMIT])
        if len(result.ambiguous_fields) > ISSUE_DETAIL_LIMIT:
            lines.append(f"      - 歧义字段过多，只列出前 {ISSUE_DETAIL_LIMIT} 条。")
    if result.type_errors:
        lines.append("   字段类型错误：")
        lines.extend(f"      - {item}" for item in result.type_errors)
    missing_total = sum(issue.count for issue in result.missing_value_issues)
    if missing_total:
        lines.append(f"   空值问题：{missing_total} 个")
        for issue in result.missing_value_issues:
            lines.append(f"      - 指标：{issue.indicator}；字段：{issue.field_name}；空值要素数：{issue.count}")
            lines.extend(format_samples(issue.samples))
    invalid_total = sum(issue.count for issue in result.invalid_category_issues)
    if invalid_total:
        lines.append(f"   类别值问题：{invalid_total} 个")
        for issue in result.invalid_category_issues:
            lines.append(
                f"      - 指标：{issue.indicator}；字段：{issue.field_name}；非法类别值：{issue.invalid_value}；要素数：{issue.count}"
            )
            allowed = "、".join(sorted(rule_set.concept_memberships[issue.indicator]))
            lines.append(f"        允许值：{allowed}")
            lines.extend(format_samples(issue.samples))
    if not lines:
        lines.append("   无。")
    return lines


def build_high_standard_preflight_text(report: HighStandardPreflightReport) -> str:
    lines: list[str] = []
    lines.append("第一步高标隶属度计算前审查报告")
    lines.append("=" * 60)
    lines.append(f"国标二级农业区：{report.rule_set.area_name}")
    lines.append(f"规则文件：{report.rule_set.rule_path}")
    lines.append(f"输入文件数：{len(report.input_sources)}")
    lines.append(f"输入要素总数：{report.feature_count}")
    lines.append(f"海拔高度是否必需：{'是' if report.rule_set.area_name in ALTITUDE_REQUIRED_AREAS else '否'}")
    lines.append("")
    lines.append("一、坐标系检查")
    lines.append(f"输入投影是否一致：{'是' if report.projections_same else '否'}")
    for info in report.projection_infos:
        lines.append(f"   - {source_label(info.source)}：{info.message}")
    lines.append(f"统一投影来源：{source_label(report.target_projection_source)}")
    lines.append("")
    lines.append("二、字段匹配和计算前检查")
    for result in report.source_results:
        lines.append(f"   文件：{source_label(result.source)}")
        lines.append(f"   要素数：{result.feature_count}")
        if result.missing_fields or result.ambiguous_fields:
            lines.extend(format_issue_block(result, report.rule_set))
            lines.append("")
            continue
        for indicator in STANDARD_INDICATOR_FIELDS:
            binding = result.field_bindings.get(indicator)
            if binding is None:
                lines.append(f"      - {indicator}: 未提供（本农业区允许缺省）")
            else:
                lines.append(f"      - {indicator} -> {binding.source_field}（{binding.field_type}）")
        lines.extend(format_issue_block(result, report.rule_set))
        lines.append("")
    lines.append("三、输出字段")
    lines.append("   输出只包含 13 个指标原值、13 个隶属度字段、评价得分、质量等级。")
    lines.append("   字段顺序按工具固定顺序创建。")
    lines.append("")
    lines.append("审查通过，可以继续计算。" if report.ok else "审查未通过，请修正输入数据后重新审查。")
    return "\n".join(lines)


def membership_for_numeric(rule: NumericRule, value: float) -> float:
    u = float(value)
    code = rule.function_code.upper()
    if code == "NEG_LINEAR_CLAMP":
        if u <= rule.lower:
            return 1.0
        if u >= rule.upper:
            return 0.0
        if rule.b is None:
            raise RuntimeError(f"{rule.indicator} 负直线型缺少 b 值。")
        return max(0.0, min(1.0, rule.b - rule.a * u))
    if code == "UPPER_SATURATION":
        if u <= rule.lower:
            return 0.0
        if u >= rule.upper:
            return 1.0
        if rule.c is None:
            raise RuntimeError(f"{rule.indicator} 戒上型缺少 c 值。")
        return max(0.0, min(1.0, 1.0 / (1.0 + rule.a * (u - rule.c) ** 2)))
    if code == "PEAK":
        if u <= rule.lower or u >= rule.upper:
            return 0.0
        if rule.c is None:
            raise RuntimeError(f"{rule.indicator} 峰型缺少 c 值。")
        return max(0.0, min(1.0, 1.0 / (1.0 + rule.a * (u - rule.c) ** 2)))
    raise RuntimeError(f"不支持的函数代码：{rule.function_code}")


def grade_for_score(score: float, grade_rules: list[GradeRule]) -> int:
    for rule in grade_rules:
        lower_ok = True if rule.lower is None else score >= rule.lower
        upper_ok = True if rule.upper is None else score < rule.upper
        if lower_ok and upper_ok:
            return rule.grade_value
    return grade_rules[-1].grade_value


def output_workspace(output_path: Path, output_feature_name: str | None) -> str:
    if output_format(output_path) == "shp":
        return str(output_path.parent)
    return str(output_path)


def validated_field_name(name: str, workspace: str, existing_upper: set[str]) -> str:
    candidate = arcpy.ValidateFieldName(name, workspace)
    base = candidate
    suffix = 1
    while candidate.upper() in existing_upper:
        suffix_text = str(suffix)
        candidate = arcpy.ValidateFieldName(f"{base[: max(1, 64 - len(suffix_text))]}{suffix_text}", workspace)
        suffix += 1
    existing_upper.add(candidate.upper())
    return candidate


def ensure_result_fields(feature_class: str, output_kind: str, rule_set: RuleSet) -> dict[str, str]:
    workspace = str(Path(feature_class).parent) if output_kind == "shp" else str(Path(feature_class).parent)
    existing = {field.name.upper(): field.name for field in arcpy.ListFields(feature_class)}
    existing_upper = set(existing)
    result_map: dict[str, str] = {}
    for indicator in rule_set.indicators:
        target = f"F{indicator}"
        actual = existing.get(target.upper())
        if actual is None:
            actual = validated_field_name(target, workspace, existing_upper)
            arcpy.management.AddField(feature_class, actual, "DOUBLE", field_alias=target)
            existing[actual.upper()] = actual
        result_map[indicator] = actual
    for result_name in (RESULT_SCORE_FIELD, RESULT_GRADE_FIELD):
        actual = existing.get(result_name.upper())
        if actual is None:
            actual = validated_field_name(result_name, workspace, existing_upper)
            field_type = "SHORT" if result_name == RESULT_GRADE_FIELD else "DOUBLE"
            arcpy.management.AddField(feature_class, actual, field_type, field_alias=result_name)
            existing[actual.upper()] = actual
        result_map[result_name] = actual
    return result_map


def copy_source_to_output(job: CalculationJob) -> str:
    delete_output_dataset(job.output_path, job.output_feature_name)
    source_dataset = source_dataset_path(job.source)
    if job.output_kind == "shp":
        job.output_path.parent.mkdir(parents=True, exist_ok=True)
        arcpy.conversion.ExportFeatures(source_dataset, str(job.output_path))
        return str(job.output_path)
    job.output_path.parent.mkdir(parents=True, exist_ok=True)
    if not job.output_path.exists():
        arcpy.management.CreateFileGDB(str(job.output_path.parent), job.output_path.name)
    if not job.output_feature_name:
        raise RuntimeError("输出到 GDB 时必须指定面要素类名称。")
    out_fc = str(job.output_path / job.output_feature_name)
    arcpy.conversion.ExportFeatures(source_dataset, out_fc)
    return out_fc


def copied_source_field_map(source_dataset: str, output_fc: str) -> dict[str, str]:
    source_fields = [field.name for field in arcpy.ListFields(source_dataset) if is_data_field(field)]
    output_fields = [field.name for field in arcpy.ListFields(output_fc) if is_data_field(field)]
    if len(output_fields) < len(source_fields):
        raise RuntimeError(
            f"输出字段数量少于输入字段数量，无法建立计算字段映射：输入 {len(source_fields)} 个，输出 {len(output_fields)} 个。"
        )
    return {source_name: output_fields[index] for index, source_name in enumerate(source_fields)}


def result_original_names(rule_set: RuleSet) -> set[str]:
    return {f"F{indicator}" for indicator in rule_set.indicators} | {RESULT_SCORE_FIELD, RESULT_GRADE_FIELD}


def source_result_field_names(source: VectorSource, rule_set: RuleSet) -> set[str]:
    names = set(result_original_names(rule_set))
    mapping = read_shp_field_mapping(source)
    for original in result_original_names(rule_set):
        actual = mapping.get(original)
        if actual:
            names.add(actual)
    return names


def delete_existing_result_fields(
    output_fc: str,
    source: VectorSource,
    rule_set: RuleSet,
    source_output_map: dict[str, str],
    logger: logging.Logger,
) -> None:
    originals = result_original_names(rule_set)
    source_result_names = source_result_field_names(source, rule_set)
    delete_names: set[str] = set()
    for source_name, output_name in source_output_map.items():
        if source_name in source_result_names:
            delete_names.add(output_name)
    for field in arcpy.ListFields(output_fc):
        if not is_data_field(field):
            continue
        alias = field.aliasName or field.name
        if field.name in originals or alias in originals or field.name in source_result_names:
            delete_names.add(field.name)
    if delete_names:
        logger.info("删除输出副本中已有计算字段：%s", sorted(delete_names))
        arcpy.management.DeleteField(output_fc, sorted(delete_names))


def bound_output_field_map(source_output_map: dict[str, str], input_bindings: dict[str, FieldBinding]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for binding in input_bindings.values():
        actual = source_output_map.get(binding.source_field)
        if actual is None:
            raise RuntimeError(f"无法在输出中找到输入字段：{binding.source_field}")
        mapping[binding.source_field] = actual
    return mapping


def copy_field_mapping_if_needed(job: CalculationJob, source_output_map: dict[str, str], result_map: dict[str, str]) -> None:
    if job.output_kind != "shp":
        return
    original_mapping = read_shp_field_mapping(job.source)
    mapping = dict(original_mapping)
    for field in arcpy.ListFields(source_dataset_path(job.source)):
        if is_data_field(field):
            mapping.setdefault(field.name, source_output_map.get(field.name, field.name))
    for indicator, actual in result_map.items():
        if indicator in {RESULT_SCORE_FIELD, RESULT_GRADE_FIELD}:
            mapping[indicator] = actual
        else:
            mapping[f"F{indicator}"] = actual
    job.output_path.with_suffix(".fields.json").write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")


def build_blank_result_report(
    feature_class: str,
    check_fields: list[str],
    title: str,
    oid_filter: set[int] | None = None,
) -> tuple[bool, str, dict[str, int]]:
    oid_name = oid_field_name(feature_class)
    excluded = {oid_name.upper(), *(field.upper() for field in check_fields)}
    sample_fields = [field for field in first_data_field_names(feature_class) if field.upper() not in excluded]
    fields = [oid_name, *check_fields, *sample_fields]
    checked = 0
    missing_by_field: dict[str, int] = {field: 0 for field in check_fields}
    issue_lines: list[str] = []
    show_details = True
    detail_count = 0
    detail_limit = ISSUE_DETAIL_LIMIT

    with arcpy.da.SearchCursor(feature_class, fields) as cursor:
        for row in cursor:
            oid = int(row[0])
            if oid_filter is not None and oid not in oid_filter:
                continue
            checked += 1
            result_values = row[1 : 1 + len(check_fields)]
            sample_values = row[1 + len(check_fields) :]
            missing_fields = [field for field, value in zip(check_fields, result_values) if is_blank_value(value)]
            if not missing_fields:
                continue
            for field in missing_fields:
                missing_by_field[field] += 1
            detail_count += 1
            if show_details and detail_count <= detail_limit:
                values = "；".join(
                    f"{name}={field_value_text(value)}" for name, value in zip(sample_fields, sample_values)
                ) or "无可展示字段值"
                issue_lines.append(f"   - OID={oid}；缺失字段：{'、'.join(missing_fields)}；{values}")
            elif detail_count > detail_limit:
                show_details = False

    missing_total = sum(missing_by_field.values())
    lines: list[str] = []
    lines.append(title)
    lines.append("=" * 60)
    lines.append(f"检查要素数：{checked}")
    lines.append(f"检查字段数：{len(check_fields)}")
    lines.append(f"缺失值总数：{missing_total}")
    lines.append("")
    if missing_total:
        lines.append("一、字段缺失统计")
        for field, count in missing_by_field.items():
            if count:
                lines.append(f"   - {field}：{count} 个要素缺失")
        lines.append("")
        lines.append("二、要素级样例")
        if detail_count > detail_limit:
            lines.append(f"缺失要素过多（{detail_count} 个），只输出字段汇总，不展开全部要素。")
            lines.extend(issue_lines)
        else:
            lines.extend(issue_lines)
    else:
        lines.append("未发现结果字段空值。")
    stats = {"checked_features": checked, "missing_values": missing_total, "issue_features": detail_count}
    return missing_total == 0, "\n".join(lines), stats


def calculate_output(job: CalculationJob, logger: logging.Logger) -> tuple[str, int, dict[str, int]]:
    require_runtime()
    arcpy.env.overwriteOutput = True
    output_fc = copy_source_to_output(job)
    copied_field_map = copied_source_field_map(source_dataset_path(job.source), output_fc)
    delete_existing_result_fields(output_fc, job.source, job.rule_set, copied_field_map, logger)
    source_output_map = bound_output_field_map(copied_field_map, job.field_bindings)
    result_map = ensure_result_fields(output_fc, job.output_kind, job.rule_set)

    read_fields = [source_output_map[binding.source_field] for binding in job.field_bindings.values()]
    indicators = list(job.field_bindings.keys())
    update_fields = read_fields + [result_map[indicator] for indicator in indicators] + [
        result_map[RESULT_SCORE_FIELD],
        result_map[RESULT_GRADE_FIELD],
    ]
    calculated = 0
    with arcpy.da.UpdateCursor(output_fc, update_fields) as cursor:
        for row in cursor:
            input_values = row[: len(indicators)]
            memberships: list[float] = []
            score = 0.0
            for indicator, value in zip(indicators, input_values):
                if indicator in job.rule_set.concept_memberships:
                    membership = job.rule_set.concept_memberships[indicator][normalize_category(value)]
                else:
                    membership = membership_for_numeric(job.rule_set.numeric_rules[indicator], float(value))
                memberships.append(membership)
                score += membership * job.rule_set.weights[indicator]
            grade = grade_for_score(score, job.rule_set.grade_rules)
            row[len(indicators) :] = memberships + [score, grade]
            cursor.updateRow(row)
            calculated += 1
    copy_field_mapping_if_needed(job, source_output_map, result_map)
    check_fields = [result_map[indicator] for indicator in indicators] + [
        result_map[RESULT_SCORE_FIELD],
        result_map[RESULT_GRADE_FIELD],
    ]
    ok, audit_text, audit_stats = build_blank_result_report(output_fc, check_fields, "隶属度计算后结果完整性审查")
    logger.info("结果完整性审计：\n%s", audit_text)
    if not ok:
        raise RuntimeError(f"计算后仍有结果字段空值：{audit_stats['missing_values']} 个。详情见日志。")
    logger.info("输出完成：%s；计算要素数：%s", output_fc, calculated)
    return output_fc, calculated, audit_stats


def output_field_type_for_indicator(rule_set: RuleSet, indicator: str) -> tuple[str, dict]:
    if indicator in INDICATOR_TEXT_FIELDS:
        return "TEXT", {"field_length": 255}
    return "DOUBLE", {}


def high_standard_expected_fields(include_overlap_field: bool) -> list[str]:
    fields = STANDARD_INDICATOR_FIELDS + [f"F{indicator}" for indicator in STANDARD_INDICATOR_FIELDS] + [
        RESULT_SCORE_FIELD,
        RESULT_GRADE_FIELD,
    ]
    if include_overlap_field:
        fields.append(OVERLAP_MARK_FIELD)
    return fields


def add_high_standard_result_fields(
    feature_class: str,
    rule_set: RuleSet,
    *,
    include_overlap_field: bool,
) -> dict[str, str]:
    field_map: dict[str, str] = {}
    for indicator in STANDARD_INDICATOR_FIELDS:
        field_type, kwargs = output_field_type_for_indicator(rule_set, indicator)
        arcpy.management.AddField(feature_class, indicator, field_type, field_alias=indicator, **kwargs)
        field_map[indicator] = indicator
    for indicator in STANDARD_INDICATOR_FIELDS:
        field_name = f"F{indicator}"
        arcpy.management.AddField(feature_class, field_name, "DOUBLE", field_alias=field_name)
        field_map[field_name] = field_name
    arcpy.management.AddField(feature_class, RESULT_SCORE_FIELD, "DOUBLE", field_alias=RESULT_SCORE_FIELD)
    arcpy.management.AddField(feature_class, RESULT_GRADE_FIELD, "SHORT", field_alias=RESULT_GRADE_FIELD)
    field_map[RESULT_SCORE_FIELD] = RESULT_SCORE_FIELD
    field_map[RESULT_GRADE_FIELD] = RESULT_GRADE_FIELD
    if include_overlap_field:
        arcpy.management.AddField(feature_class, OVERLAP_MARK_FIELD, "SHORT", field_alias=OVERLAP_MARK_FIELD)
        field_map[OVERLAP_MARK_FIELD] = OVERLAP_MARK_FIELD
    return field_map


def create_high_standard_output_feature_class(job: HighStandardCalculationJob) -> tuple[str, dict[str, str]]:
    if job.output_kind != "gdb":
        raise RuntimeError("第一步固定输出为 GDB 面要素类，以保留完整中文字段名。")
    delete_output_dataset(job.output_path, job.output_feature_name)
    job.output_path.parent.mkdir(parents=True, exist_ok=True)
    if not job.output_path.exists():
        arcpy.management.CreateFileGDB(str(job.output_path.parent), job.output_path.name)
    if not job.output_feature_name:
        raise RuntimeError("输出到 GDB 时必须指定面要素类名称。")
    output_fc = str(job.output_path / job.output_feature_name)
    arcpy.management.CreateFeatureclass(str(job.output_path), job.output_feature_name, "POLYGON", spatial_reference=job.target_spatial_reference)
    include_overlap_field = job.overlap_mode == OVERLAP_MODE_MARK
    field_map = add_high_standard_result_fields(output_fc, job.rule_set, include_overlap_field=include_overlap_field)
    expected_fields = high_standard_expected_fields(include_overlap_field)
    actual_fields = data_field_names(output_fc)
    if actual_fields != expected_fields:
        missing = [name for name in expected_fields if name not in actual_fields]
        extra = [name for name in actual_fields if name not in expected_fields]
        raise RuntimeError(
            "第一步输出字段结构不符合固定清单。"
            f"缺少：{'、'.join(missing) or '无'}；"
            f"多余：{'、'.join(extra) or '无'}；"
            f"实际顺序：{'、'.join(actual_fields)}"
        )
    return output_fc, field_map


def field_value_signature(value: object) -> object:
    if is_blank_value(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip()


def geometry_signature(geometry: object) -> bytes | str:
    try:
        return bytes(geometry.WKB)
    except Exception:
        return getattr(geometry, "JSON", str(geometry))


def read_high_standard_candidates(
    job: HighStandardCalculationJob,
    temp_dir: Path,
    logger: logging.Logger,
) -> tuple[list[HighStandardFeatureCandidate], int]:
    prepared_sources = vector_tool.prepare_sources(job.input_sources, job.target_spatial_reference, temp_dir, logger)
    result_by_source = {source_label(result.source): result for result in job.source_results}
    candidates: list[HighStandardFeatureCandidate] = []
    duplicate_index: dict[tuple[bytes | str, tuple[object, ...]], HighStandardFeatureCandidate] = {}
    duplicate_count = 0

    for source_order, prepared in enumerate(prepared_sources, start=1):
        source_result = result_by_source[source_label(prepared.source)]
        bound_indicators = [indicator for indicator in STANDARD_INDICATOR_FIELDS if indicator in source_result.field_bindings]
        read_fields = ["OID@", "SHAPE@", *[source_result.field_bindings[indicator].source_field for indicator in bound_indicators]]
        with arcpy.da.SearchCursor(prepared.path, read_fields) as search_cursor:
            for row in search_cursor:
                source_oid = int(row[0])
                geometry = row[1]
                values_by_indicator = {indicator: value for indicator, value in zip(bound_indicators, row[2:])}
                memberships_by_indicator: dict[str, float | None] = {}
                score = 0.0
                for indicator in job.rule_set.indicators:
                    if indicator not in values_by_indicator:
                        memberships_by_indicator[indicator] = None
                        continue
                    value = values_by_indicator[indicator]
                    if indicator in job.rule_set.concept_memberships:
                        membership = job.rule_set.concept_memberships[indicator][normalize_category(value)]
                    else:
                        membership = membership_for_numeric(job.rule_set.numeric_rules[indicator], float(value))
                    memberships_by_indicator[indicator] = membership
                    score += membership * job.rule_set.weights[indicator]
                grade = grade_for_score(score, job.rule_set.grade_rules)
                raw_values = [values_by_indicator.get(indicator) for indicator in STANDARD_INDICATOR_FIELDS]
                duplicate_key = (
                    geometry_signature(geometry),
                    tuple(field_value_signature(value) for value in raw_values),
                )
                existing = duplicate_index.get(duplicate_key)
                if existing is not None and existing.source_label != source_label(prepared.source):
                    duplicate_count += 1
                    logger.info(
                        "完全相同要素自动去重：保留 %s OID=%s，跳过 %s OID=%s",
                        existing.source_label,
                        existing.source_oid,
                        source_label(prepared.source),
                        source_oid,
                    )
                    continue
                candidate = HighStandardFeatureCandidate(
                    candidate_id=len(candidates) + 1,
                    source_label=source_label(prepared.source),
                    source_order=source_order,
                    source_oid=source_oid,
                    geometry=geometry,
                    raw_values=raw_values,
                    membership_values=[memberships_by_indicator.get(indicator) for indicator in STANDARD_INDICATOR_FIELDS],
                    score=score,
                    grade=grade,
                )
                candidates.append(candidate)
                duplicate_index.setdefault(duplicate_key, candidate)

    if duplicate_count:
        logger.info("完全相同几何和 13 个评价指标的跨来源要素已自动去重：%s 个。", duplicate_count)
    return candidates, duplicate_count


def create_candidate_index_feature_class(
    temp_dir: Path,
    name: str,
    spatial_reference: object,
    candidates: list[HighStandardFeatureCandidate],
) -> str:
    gdb_path = temp_dir / f"{name}.gdb"
    arcpy.management.CreateFileGDB(str(temp_dir), gdb_path.name)
    feature_class = str(gdb_path / "features")
    arcpy.management.CreateFeatureclass(str(gdb_path), "features", "POLYGON", spatial_reference=spatial_reference)
    arcpy.management.AddField(feature_class, TEMP_CANDIDATE_ID_FIELD, "LONG")
    arcpy.management.AddField(feature_class, TEMP_SOURCE_ORDER_FIELD, "LONG")
    arcpy.management.AddField(feature_class, TEMP_SOURCE_OID_FIELD, "LONG")
    fields = ["SHAPE@", TEMP_CANDIDATE_ID_FIELD, TEMP_SOURCE_ORDER_FIELD, TEMP_SOURCE_OID_FIELD]
    with arcpy.da.InsertCursor(feature_class, fields) as cursor:
        for candidate in candidates:
            cursor.insertRow([candidate.geometry, candidate.candidate_id, candidate.source_order, candidate.source_oid])
    return feature_class


def suffixed_field_names(feature_class: str, base_name: str) -> list[str]:
    base = base_name.upper()
    return [field.name for field in arcpy.ListFields(feature_class) if field.name.upper() == base or field.name.upper().startswith(f"{base}_")]


def detect_candidate_overlaps(
    candidates: list[HighStandardFeatureCandidate],
    job: HighStandardCalculationJob,
    temp_dir: Path,
    logger: logging.Logger,
) -> tuple[set[int], int]:
    if len(candidates) < 2:
        return set(), 0
    candidate_by_id = {candidate.candidate_id: candidate for candidate in candidates}
    index_fc = create_candidate_index_feature_class(temp_dir, "step1_overlap_a", job.target_spatial_reference, candidates)
    copy_fc = str(temp_dir / "step1_overlap_b.gdb" / "features")
    arcpy.management.CreateFileGDB(str(temp_dir), "step1_overlap_b.gdb")
    arcpy.management.CopyFeatures(index_fc, copy_fc)
    intersect_gdb = temp_dir / "step1_overlap_intersect.gdb"
    arcpy.management.CreateFileGDB(str(temp_dir), intersect_gdb.name)
    intersect_fc = str(intersect_gdb / "overlap")
    arcpy.analysis.PairwiseIntersect([index_fc, copy_fc], intersect_fc, "ALL")

    id_fields = suffixed_field_names(intersect_fc, TEMP_CANDIDATE_ID_FIELD)
    source_fields = suffixed_field_names(intersect_fc, TEMP_SOURCE_ORDER_FIELD)
    if len(id_fields) < 2 or len(source_fields) < 2:
        logger.warning("重叠检测结果字段异常，无法输出重叠明细字段名：%s", [field.name for field in arcpy.ListFields(intersect_fc)])
        return set(), 0

    overlap_ids: set[int] = set()
    seen_pairs: set[tuple[int, int]] = set()
    detail_count = 0
    detail_limit = 200
    with arcpy.da.SearchCursor(intersect_fc, [id_fields[0], id_fields[1], source_fields[0], source_fields[1], "SHAPE@AREA"]) as cursor:
        for id_a, id_b, source_a, source_b, area in cursor:
            if id_a == id_b or source_a == source_b or not area or area <= 0:
                continue
            pair = tuple(sorted((int(id_a), int(id_b))))
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            overlap_ids.update(pair)
            if detail_count < detail_limit:
                first = candidate_by_id[pair[0]]
                second = candidate_by_id[pair[1]]
                logger.warning(
                    "非完全相同跨来源重叠：面积=%.6f；%s OID=%s；%s OID=%s",
                    float(area),
                    first.source_label,
                    first.source_oid,
                    second.source_label,
                    second.source_oid,
                )
                detail_count += 1
    if len(seen_pairs) > detail_limit:
        logger.warning("非完全相同跨来源重叠关系过多，共 %s 组；日志仅列出前 %s 组。", len(seen_pairs), detail_limit)
    if seen_pairs:
        logger.warning(
            "检测到非完全相同跨来源重叠：涉及 %s 个候选要素、%s 组重叠关系。",
            len(overlap_ids),
            len(seen_pairs),
        )
    return overlap_ids, len(seen_pairs)


def create_calculated_temp_feature_class(
    temp_dir: Path,
    name: str,
    spatial_reference: object,
    rule_set: RuleSet,
) -> tuple[str, dict[str, str]]:
    gdb_path = temp_dir / f"{name}.gdb"
    arcpy.management.CreateFileGDB(str(temp_dir), gdb_path.name)
    feature_class = str(gdb_path / "features")
    arcpy.management.CreateFeatureclass(str(gdb_path), "features", "POLYGON", spatial_reference=spatial_reference)
    arcpy.management.AddField(feature_class, TEMP_CANDIDATE_ID_FIELD, "LONG")
    field_map = add_high_standard_result_fields(feature_class, rule_set, include_overlap_field=False)
    return feature_class, field_map


def insert_candidates_to_feature_class(
    feature_class: str,
    field_map: dict[str, str],
    candidates: list[HighStandardFeatureCandidate],
) -> None:
    fields = [
        "SHAPE@",
        TEMP_CANDIDATE_ID_FIELD,
        *[field_map[indicator] for indicator in STANDARD_INDICATOR_FIELDS],
        *[field_map[f"F{indicator}"] for indicator in STANDARD_INDICATOR_FIELDS],
        field_map[RESULT_SCORE_FIELD],
        field_map[RESULT_GRADE_FIELD],
    ]
    with arcpy.da.InsertCursor(feature_class, fields) as cursor:
        for candidate in candidates:
            cursor.insertRow([
                candidate.geometry,
                candidate.candidate_id,
                *candidate.raw_values,
                *candidate.membership_values,
                candidate.score,
                candidate.grade,
            ])


def append_candidate_rows(
    source_fc: str,
    output_fc: str,
    field_map: dict[str, str],
    *,
    overlap_ids: set[int] | None = None,
) -> int:
    include_overlap = overlap_ids is not None and OVERLAP_MARK_FIELD in field_map
    read_fields = [
        "SHAPE@",
        TEMP_CANDIDATE_ID_FIELD,
        *[indicator for indicator in STANDARD_INDICATOR_FIELDS],
        *[f"F{indicator}" for indicator in STANDARD_INDICATOR_FIELDS],
        RESULT_SCORE_FIELD,
        RESULT_GRADE_FIELD,
    ]
    insert_fields = [
        "SHAPE@",
        *[field_map[indicator] for indicator in STANDARD_INDICATOR_FIELDS],
        *[field_map[f"F{indicator}"] for indicator in STANDARD_INDICATOR_FIELDS],
        field_map[RESULT_SCORE_FIELD],
        field_map[RESULT_GRADE_FIELD],
    ]
    if include_overlap:
        insert_fields.append(field_map[OVERLAP_MARK_FIELD])
    inserted = 0
    with arcpy.da.SearchCursor(source_fc, read_fields) as search_cursor, arcpy.da.InsertCursor(output_fc, insert_fields) as insert_cursor:
        for row in search_cursor:
            candidate_id = int(row[1])
            values = list(row[2:])
            if include_overlap:
                values.append(1 if candidate_id in overlap_ids else 0)
            insert_cursor.insertRow([row[0], *values])
            inserted += 1
    return inserted


def insert_marked_candidates(
    output_fc: str,
    field_map: dict[str, str],
    candidates: list[HighStandardFeatureCandidate],
    overlap_ids: set[int],
) -> int:
    fields = [
        "SHAPE@",
        *[field_map[indicator] for indicator in STANDARD_INDICATOR_FIELDS],
        *[field_map[f"F{indicator}"] for indicator in STANDARD_INDICATOR_FIELDS],
        field_map[RESULT_SCORE_FIELD],
        field_map[RESULT_GRADE_FIELD],
        field_map[OVERLAP_MARK_FIELD],
    ]
    with arcpy.da.InsertCursor(output_fc, fields) as cursor:
        for candidate in candidates:
            cursor.insertRow([
                candidate.geometry,
                *candidate.raw_values,
                *candidate.membership_values,
                candidate.score,
                candidate.grade,
                1 if candidate.candidate_id in overlap_ids else 0,
            ])
    return len(candidates)


def insert_order_erased_candidates(
    output_fc: str,
    field_map: dict[str, str],
    candidates: list[HighStandardFeatureCandidate],
    job: HighStandardCalculationJob,
    temp_dir: Path,
    logger: logging.Logger,
) -> int:
    kept_fc = None
    inserted = 0
    for source_order in sorted({candidate.source_order for candidate in candidates}):
        group = [candidate for candidate in candidates if candidate.source_order == source_order]
        if not group:
            continue
        group_fc, _group_map = create_calculated_temp_feature_class(temp_dir, f"step1_order_{source_order:03d}", job.target_spatial_reference, job.rule_set)
        insert_candidates_to_feature_class(group_fc, _group_map, group)
        effective_fc = group_fc
        if kept_fc is not None and int(arcpy.management.GetCount(kept_fc)[0]) > 0:
            erased_gdb = temp_dir / f"step1_erased_{source_order:03d}.gdb"
            arcpy.management.CreateFileGDB(str(temp_dir), erased_gdb.name)
            effective_fc = str(erased_gdb / "features")
            arcpy.analysis.Erase(group_fc, kept_fc, effective_fc)
            logger.info("按输入顺序去重：第 %s 个数据源被前序保留区域 Erase。", source_order)
        inserted += append_candidate_rows(effective_fc, output_fc, field_map)
        if kept_fc is None:
            kept_gdb = temp_dir / "step1_kept_union.gdb"
            arcpy.management.CreateFileGDB(str(temp_dir), kept_gdb.name)
            kept_fc = str(kept_gdb / "features")
            arcpy.management.CopyFeatures(effective_fc, kept_fc)
        else:
            arcpy.management.Append(effective_fc, kept_fc, "NO_TEST")
    return inserted


def calculate_high_standard_output(
    job: HighStandardCalculationJob,
    temp_dir: Path,
    logger: logging.Logger,
) -> tuple[str, int, dict[str, int]]:
    require_runtime()
    arcpy.env.overwriteOutput = True
    output_fc, field_map = create_high_standard_output_feature_class(job)
    candidates, duplicate_count = read_high_standard_candidates(job, temp_dir, logger)
    overlap_ids: set[int] = set()
    overlap_pair_count = 0
    if candidates:
        overlap_ids, overlap_pair_count = detect_candidate_overlaps(candidates, job, temp_dir, logger)
    logger.info(
        "候选要素统计：保留候选 %s 个；完全相同跨来源要素自动去重 %s 个；非完全相同重叠关系 %s 组。",
        len(candidates),
        duplicate_count,
        overlap_pair_count,
    )
    if job.overlap_mode == OVERLAP_MODE_MARK:
        calculated = insert_marked_candidates(output_fc, field_map, candidates, overlap_ids)
        if overlap_ids:
            logger.warning("已在输出字段“%s”中用 1 标记 %s 个存在非完全相同跨来源重叠的要素。", OVERLAP_MARK_FIELD, len(overlap_ids))
    elif job.overlap_mode == OVERLAP_MODE_ORDER_ERASE:
        calculated = insert_order_erased_candidates(output_fc, field_map, candidates, job, temp_dir, logger)
    else:
        raise RuntimeError(f"不支持的重叠处理方式：{job.overlap_mode}")

    optional = optional_indicators_for_area(job.rule_set.area_name)
    required_indicators = [
        indicator
        for indicator in STANDARD_INDICATOR_FIELDS
        if indicator not in optional or any(indicator in result.field_bindings for result in job.source_results)
    ]
    check_fields = [field_map[indicator] for indicator in required_indicators]
    check_fields += [field_map[f"F{indicator}"] for indicator in required_indicators if indicator in job.rule_set.indicators]
    check_fields += [field_map[RESULT_SCORE_FIELD], field_map[RESULT_GRADE_FIELD]]
    if job.overlap_mode == OVERLAP_MODE_MARK:
        check_fields += [field_map[OVERLAP_MARK_FIELD]]
    ok, audit_text, audit_stats = build_blank_result_report(output_fc, check_fields, "第一步高标隶属度计算后结果完整性审查")
    logger.info("结果完整性审计：\n%s", audit_text)
    if not ok:
        raise RuntimeError(f"计算后仍有结果字段空值：{audit_stats['missing_values']} 个。详情见日志。")
    logger.info("输出完成：%s；计算要素数：%s", output_fc, calculated)
    return output_fc, calculated, audit_stats


def setup_job_logger(logs_dir: Path, job_id: str) -> tuple[logging.Logger, Path]:
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / f"membership_{timestamp_for_file()}_{job_id}.log"
    logger = logging.getLogger(f"membership_arcpy.{job_id}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(log_path, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(handler)
    return logger, log_path


def close_job_logger(logger: logging.Logger) -> None:
    for handler in list(logger.handlers):
        handler.close()
        logger.removeHandler(handler)


def append_history(history_path: Path, record: dict) -> None:
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(record, ensure_ascii=False) + "\n")


class CalculationWorker(threading.Thread):
    def __init__(self, job_queue: queue.Queue, event_queue: queue.Queue, logs_dir: Path, process_dir: Path):
        super().__init__(daemon=True)
        self.job_queue = job_queue
        self.event_queue = event_queue
        self.logs_dir = logs_dir
        self.process_dir = process_dir
        self.history_path = logs_dir / "membership_history.jsonl"

    def send(self, event_type: str, payload: dict) -> None:
        self.event_queue.put((event_type, payload))

    def run(self) -> None:
        while True:
            job = self.job_queue.get()
            if job is None:
                self.job_queue.task_done()
                break
            self.process_job(job)
            self.job_queue.task_done()

    def process_job(self, job: HighStandardCalculationJob) -> None:
        logger, log_path = setup_job_logger(self.logs_dir, job.job_id)
        temp_dir = self.process_dir / f"membership_{timestamp_for_file()}_{job.job_id}"
        output_target = str(job.output_path / job.output_feature_name) if job.output_kind == "gdb" else str(job.output_path)
        record = {
            "job_id": job.job_id,
            "created_at": job.created_at,
            "started_at": now_text(),
            "ended_at": None,
            "status": "running",
            "input_sources": [source_label(source) for source in job.input_sources],
            "output_path": output_target,
            "overlap_mode": job.overlap_mode,
            "area_name": job.rule_set.area_name,
            "rule_path": str(job.rule_set.rule_path),
            "target_projection_source": source_label(job.target_projection_source),
            "target_projection": vector_tool.describe_spatial_reference(job.target_spatial_reference),
            "validation_report": job.validation_report,
            "log_path": str(log_path),
            "error": None,
            "calculated_count": None,
            "audit_stats": None,
        }
        self.send("job_started", {"job_id": job.job_id, "message": "开始计算高标隶属度", "log_path": str(log_path)})
        try:
            require_runtime()
            temp_dir.mkdir(parents=True, exist_ok=True)
            logger.info("任务开始：%s", job.job_id)
            logger.info("输入源：%s", [source_label(source) for source in job.input_sources])
            logger.info("输出目标：%s", output_target)
            logger.info("统一投影：%s", vector_tool.describe_spatial_reference(job.target_spatial_reference))
            logger.info("规则文件：%s", job.rule_set.rule_path)
            logger.info("重叠处理方式：%s", job.overlap_mode)
            logger.info("检查报告：\n%s", job.validation_report)
            output_fc, count, audit_stats = calculate_high_standard_output(job, temp_dir, logger)
            record.update({"status": "success", "calculated_count": count, "output_path": output_fc, "audit_stats": audit_stats})
            self.send(
                "job_done",
                {
                    "job_id": job.job_id,
                    "message": f"高标隶属度计算完成：{output_fc}；要素数 {count}",
                    "output_path": output_fc,
                    "log_path": str(log_path),
                },
            )
        except Exception as exc:
            error_text = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
            logger.error(error_text)
            record.update({"status": "failed", "error": str(exc)})
            self.send(
                "job_failed",
                {
                    "job_id": job.job_id,
                    "message": f"计算失败：{exc}",
                    "log_path": str(log_path),
                },
            )
        finally:
            record["ended_at"] = now_text()
            append_history(self.history_path, record)
            shutil.rmtree(temp_dir, ignore_errors=True)
            logger.info("过程目录已清理：%s", temp_dir)
            close_job_logger(logger)


class MembershipApp:
    def __init__(
        self,
        root: Tk,
        *,
        embedded: bool = False,
        shared_status_text: Text | None = None,
        on_job_done: Callable[[dict], None] | None = None,
    ):
        self.root = root
        self.embedded = embedded
        self.shared_status_text = shared_status_text
        self.on_job_done = on_job_done
        self.paths = resolve_paths(Path.cwd())
        self.rules_dir = self.paths.data_dir / "rules"
        self.logs_dir = self.paths.outputs_dir / "logs"
        self.process_dir = self.paths.outputs_dir / "process_files"
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        self.process_dir.mkdir(parents=True, exist_ok=True)
        self.area_options = self.load_area_options()

        self.area_var = StringVar(value=self.area_options[0] if self.area_options else DEFAULT_AREA_NAME)
        self.folder_var = StringVar()
        self.reference_mode = IntVar(value=0)
        self.reference_file_var = StringVar()
        self.reference_extra_var = StringVar()
        self.output_gdb_var = StringVar()
        self.output_feature_var = StringVar(value="Step1_高标隶属度")
        self.overlap_mode_var = StringVar(value=OVERLAP_MODE_MARK)
        self.discovered_sources: list[VectorSource] = []
        self.analyzed_sources: tuple[VectorSource, ...] = ()
        self.projections_same = False
        self.projection_infos: list[object] = []
        self.target_spatial_reference = None
        self.target_source: VectorSource | None = None
        self.last_report: HighStandardPreflightReport | None = None
        self.last_report_key: tuple | None = None

        self.job_queue: queue.Queue = queue.Queue()
        self.event_queue: queue.Queue = queue.Queue()
        self.worker = CalculationWorker(self.job_queue, self.event_queue, self.logs_dir, self.process_dir)
        self.worker.start()

        if not self.embedded:
            self.root.title("第一步：计算高标隶属度（ArcPy）")
            self.root.geometry("1180x860")
            self.root.minsize(1180, 780)
        self.build_ui()
        self.root.after(200, self.poll_worker_events)

    def load_area_options(self) -> list[str]:
        options = list_rule_area_options(self.rules_dir)
        if DEFAULT_AREA_NAME in options:
            options.remove(DEFAULT_AREA_NAME)
            options.insert(0, DEFAULT_AREA_NAME)
        return options

    def refresh_area_options(self) -> None:
        self.area_options = self.load_area_options()
        self.area_combo["values"] = self.area_options
        if not self.area_options:
            self.area_var.set("")
            self.log_status(f"没有在规则目录中找到规则文件：{self.rules_dir}")
            messagebox.showwarning("提示", f"没有在规则目录中找到 *_机器读取规则.xlsx：\n{self.rules_dir}")
            return
        if self.area_var.get() not in self.area_options:
            self.area_var.set(self.area_options[0])
        self.last_report = None
        self.log_status(f"已刷新国标二级农业区规则列表：{len(self.area_options)} 个。")

    def build_ui(self) -> None:
        container = ttk.Frame(self.root, padding=10)
        container.pack(fill=BOTH, expand=True)

        input_frame = ttk.LabelFrame(container, text="1. 输入数据和规则")
        input_frame.pack(fill="x", pady=5)
        area_row = ttk.Frame(input_frame)
        area_row.pack(fill="x", padx=5, pady=5)
        ttk.Label(area_row, text="国标二级农业区").pack(side=LEFT)
        self.area_combo = ttk.Combobox(area_row, textvariable=self.area_var, state="readonly", values=self.area_options, width=36)
        self.area_combo.pack(side=LEFT, padx=5)
        ttk.Button(area_row, text="检查规则文件", command=self.check_rule_file).pack(side=LEFT, padx=5)
        ttk.Button(area_row, text="刷新规则列表", command=self.refresh_area_options).pack(side=LEFT, padx=5)

        source_frame = ttk.LabelFrame(container, text="2. 添加高标准农田面数据")
        source_frame.pack(fill="x", pady=5)
        source_row = ttk.Frame(source_frame)
        source_row.pack(fill="x", padx=5, pady=5)
        ttk.Entry(source_row, textvariable=self.folder_var).pack(side=LEFT, fill="x", expand=True, padx=5)
        ttk.Button(source_row, text="添加文件夹", command=self.choose_folder).pack(side=LEFT, padx=3)
        ttk.Button(source_row, text="添加 shp", command=self.add_shp_files).pack(side=LEFT, padx=3)
        ttk.Button(source_row, text="添加 gdb", command=self.add_gdb_folder).pack(side=LEFT, padx=3)
        ttk.Button(source_row, text="刷新清单", command=self.refresh_sources).pack(side=LEFT, padx=3)

        file_frame = ttk.LabelFrame(container, text="3. 勾选需要计算的面数据")
        file_frame.pack(fill="x", pady=5)
        buttons = ttk.Frame(file_frame)
        buttons.pack(fill="x")
        ttk.Button(buttons, text="全选", command=lambda: self.source_checklist.select_all(True)).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(buttons, text="全不选", command=lambda: self.source_checklist.select_all(False)).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(buttons, text="审查字段和投影", command=self.validate_current_input).pack(side=LEFT, padx=5, pady=4)
        self.source_checklist = vector_tool.ScrollableCheckList(file_frame)
        self.source_checklist.pack(fill="both", expand=True, padx=5, pady=5)

        projection_frame = ttk.LabelFrame(container, text="4. 投影确认")
        projection_frame.pack(fill="both", pady=5)
        projection_table = ttk.Frame(projection_frame)
        projection_table.pack(fill="x", padx=5, pady=5)
        self.projection_tree = ttk.Treeview(projection_table, columns=("file", "projection"), show="headings", height=5)
        self.projection_tree.heading("file", text="数据源")
        self.projection_tree.heading("projection", text="投影")
        self.projection_tree.column("file", width=1300, anchor="w", stretch=False)
        self.projection_tree.column("projection", width=800, anchor="w", stretch=False)
        projection_y = ttk.Scrollbar(projection_table, orient="vertical", command=self.projection_tree.yview)
        projection_x = ttk.Scrollbar(projection_table, orient="horizontal", command=self.projection_tree.xview)
        self.projection_tree.configure(yscrollcommand=projection_y.set, xscrollcommand=projection_x.set)
        self.projection_tree.grid(row=0, column=0, sticky="nsew")
        projection_y.grid(row=0, column=1, sticky="ns")
        projection_x.grid(row=1, column=0, sticky="ew")
        projection_table.columnconfigure(0, weight=1)

        reference_frame = ttk.Frame(projection_frame)
        reference_frame.pack(fill="x", padx=5, pady=2)
        ttk.Radiobutton(reference_frame, text="使用选中的数据源作为基准", variable=self.reference_mode, value=0, command=self.update_target_projection).pack(side=LEFT)
        self.reference_combo = ttk.Combobox(reference_frame, textvariable=self.reference_file_var, state="readonly", width=72)
        self.reference_combo.pack(side=LEFT, padx=5)
        self.reference_combo.bind("<<ComboboxSelected>>", lambda _event: self.update_target_projection())
        ttk.Radiobutton(reference_frame, text="使用额外参考文件", variable=self.reference_mode, value=1, command=self.update_target_projection).pack(side=LEFT, padx=5)
        ttk.Button(reference_frame, text="选择参考 shp", command=self.choose_extra_reference).pack(side=LEFT)

        projection_text_frame = ttk.Frame(projection_frame)
        projection_text_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.projection_text = Text(projection_text_frame, height=6, wrap="none")
        projection_text_y = ttk.Scrollbar(projection_text_frame, orient="vertical", command=self.projection_text.yview)
        projection_text_x = ttk.Scrollbar(projection_text_frame, orient="horizontal", command=self.projection_text.xview)
        self.projection_text.configure(yscrollcommand=projection_text_y.set, xscrollcommand=projection_text_x.set)
        self.projection_text.grid(row=0, column=0, sticky="nsew")
        projection_text_y.grid(row=0, column=1, sticky="ns")
        projection_text_x.grid(row=1, column=0, sticky="ew")
        projection_text_frame.rowconfigure(0, weight=1)
        projection_text_frame.columnconfigure(0, weight=1)
        self.projection_text.insert(END, "选择并审查输入数据后，这里会显示统一投影。")

        overlap_frame = ttk.LabelFrame(container, text="5. 源间重叠处理")
        overlap_frame.pack(fill="x", pady=5)
        ttk.Radiobutton(
            overlap_frame,
            text="标记重叠并输出明细（推荐先用）",
            variable=self.overlap_mode_var,
            value=OVERLAP_MODE_MARK,
        ).pack(anchor="w", padx=5, pady=2)
        ttk.Radiobutton(
            overlap_frame,
            text="按输入顺序 Erase，前面的数据源优先保留",
            variable=self.overlap_mode_var,
            value=OVERLAP_MODE_ORDER_ERASE,
        ).pack(anchor="w", padx=5, pady=2)

        output_frame = ttk.LabelFrame(container, text="6. 输出位置")
        output_frame.pack(fill="x", pady=5)
        kind_row = ttk.Frame(output_frame)
        kind_row.pack(fill="x", padx=5, pady=2)
        ttk.Label(kind_row, text="第一步固定输出为 GDB 面要素类，以保留完整中文字段名和固定字段顺序。").pack(side=LEFT)

        gdb_row = ttk.Frame(output_frame)
        gdb_row.pack(fill="x", padx=5, pady=2)
        ttk.Label(gdb_row, text="GDB").pack(side=LEFT)
        ttk.Entry(gdb_row, textvariable=self.output_gdb_var).pack(side=LEFT, fill="x", expand=True, padx=5)
        ttk.Button(gdb_row, text="选择已有 gdb", command=self.choose_existing_output_gdb).pack(side=LEFT, padx=3)
        ttk.Button(gdb_row, text="新建 gdb", command=self.choose_output_gdb).pack(side=LEFT, padx=3)
        ttk.Label(gdb_row, text="面要素类").pack(side=LEFT, padx=5)
        ttk.Entry(gdb_row, textvariable=self.output_feature_var, width=24).pack(side=LEFT)

        report_frame = ttk.LabelFrame(container, text="7. 审查报告、日志和历史记录")
        report_frame.pack(fill="both", expand=True, pady=5)
        action_row = ttk.Frame(report_frame)
        action_row.pack(fill="x")
        ttk.Button(action_row, text="开始审查", command=self.validate_current_input).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(action_row, text="提交计算任务", command=self.submit_job).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(action_row, text="查看历史记录", command=self.show_history).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(action_row, text="打开日志文件夹", command=self.open_logs_folder).pack(side=LEFT, padx=5, pady=4)
        ttk.Button(action_row, text="删除日志", command=self.delete_logs).pack(side=LEFT, padx=5, pady=4)
        if self.shared_status_text is None:
            self.status_text = Text(report_frame, height=22, wrap="word")
            self.status_text.pack(fill="both", expand=True, padx=5, pady=5)
        else:
            self.status_text = self.shared_status_text
            ttk.Label(report_frame, text="运行详细信息显示在窗口底部“详细信息”区域。").pack(anchor="w", padx=5, pady=5)
        self.log_status("第一步工具已启动。请添加高标准农田面数据，选择国标二级农业区后执行审查。")

    def check_rule_file(self) -> None:
        if not self.area_var.get().strip():
            messagebox.showwarning("提示", "请先选择国标二级农业区。")
            return
        try:
            rule_set = load_rule_set(self.rules_dir, self.area_var.get())
        except Exception as exc:
            messagebox.showerror("规则文件错误", str(exc))
            return
        messagebox.showinfo("规则文件", f"已找到并读取规则文件：\n{rule_set.rule_path}\n\n指标数：{len(rule_set.weights)}")
        self.log_status(f"规则文件检查通过：{rule_set.rule_path}")

    def choose_folder(self) -> None:
        folder = filedialog.askdirectory(title="选择一个可添加数据源的文件夹")
        if folder:
            self.folder_var.set(folder)
            self.add_folder_sources(Path(folder))

    def add_folder_sources(self, folder: Path) -> None:
        if not folder.exists():
            messagebox.showwarning("提示", "请先选择有效文件夹。")
            return
        sources, errors = vector_tool.discover_sources_in_folder_with_errors(folder)
        before = len(self.discovered_sources)
        self.discovered_sources = vector_tool.merge_source_lists(self.discovered_sources, sources)
        self.discovered_sources.sort(key=vector_tool.source_path_for_log)
        self.source_checklist.set_sources(self.discovered_sources)
        self.last_report = None
        self.log_status(f"已添加 {len(self.discovered_sources) - before} 个面数据源，当前清单 {len(self.discovered_sources)} 个。")
        for error in errors:
            self.log_status(f"跳过 gdb：{error}")

    def add_shp_files(self) -> None:
        paths = filedialog.askopenfilenames(title="选择输入 shp", filetypes=[("Shapefile", "*.shp")])
        if not paths:
            return
        sources = [make_vector_source("shp", Path(path)) for path in paths]
        before = len(self.discovered_sources)
        self.discovered_sources = vector_tool.merge_source_lists(self.discovered_sources, sources)
        self.discovered_sources.sort(key=vector_tool.source_path_for_log)
        self.source_checklist.set_sources(self.discovered_sources)
        self.last_report = None
        self.log_status(f"已添加 {len(self.discovered_sources) - before} 个 shp，当前清单共 {len(self.discovered_sources)} 个。")

    def add_gdb_folder(self) -> None:
        path = filedialog.askdirectory(title="选择输入 .gdb 文件夹")
        if not path:
            return
        gdb_path = find_nearest_gdb_path(Path(path))
        if gdb_path is None or not is_gdb_path(gdb_path):
            messagebox.showwarning("提示", "请选择一个 .gdb 文件夹。")
            return
        try:
            sources = vector_tool.list_gdb_polygon_layers(gdb_path)
        except Exception as exc:
            messagebox.showerror("读取失败", str(exc))
            return
        if not sources:
            messagebox.showwarning("提示", "这个 gdb 中没有识别到面图层。")
            return
        before = len(self.discovered_sources)
        self.discovered_sources = vector_tool.merge_source_lists(self.discovered_sources, sources)
        self.discovered_sources.sort(key=vector_tool.source_path_for_log)
        self.source_checklist.set_sources(self.discovered_sources)
        self.last_report = None
        self.log_status(f"已添加 {len(self.discovered_sources) - before} 个 gdb 面图层，当前清单 {len(self.discovered_sources)} 个。")

    def refresh_sources(self) -> None:
        folder_text = self.folder_var.get().strip()
        if not folder_text:
            messagebox.showwarning("提示", "请先选择或输入一个文件夹。")
            return
        folder = Path(folder_text)
        if not folder.exists():
            messagebox.showwarning("提示", "请先选择有效文件夹。")
            return
        self.discovered_sources, errors = vector_tool.discover_sources_in_folder_with_errors(folder)
        self.discovered_sources.sort(key=vector_tool.source_path_for_log)
        self.source_checklist.set_sources(self.discovered_sources)
        self.last_report = None
        self.log_status(f"发现 {len(self.discovered_sources)} 个可用面数据源。")
        for error in errors:
            self.log_status(f"跳过 gdb：{error}")

    def choose_output_gdb(self) -> None:
        path = filedialog.asksaveasfilename(title="选择或新建 FileGDB", defaultextension=".gdb", filetypes=[("File Geodatabase", "*.gdb")])
        if path:
            if not path.lower().endswith(".gdb"):
                path += ".gdb"
            self.output_gdb_var.set(path)

    def choose_existing_output_gdb(self) -> None:
        path = filedialog.askdirectory(title="选择已有 .gdb 文件夹")
        if not path:
            return
        gdb_path = find_nearest_gdb_path(Path(path))
        if gdb_path is None or not is_gdb_path(gdb_path):
            messagebox.showwarning("提示", "请选择一个已有的 .gdb 文件夹。")
            return
        self.output_gdb_var.set(str(gdb_path))
        self.log_status(f"已选择已有 GDB 输出库：{gdb_path}")

    def selected_sources(self) -> list[VectorSource]:
        return self.source_checklist.selected_sources()

    def refresh_reference_choices(self, sources: list[VectorSource]) -> None:
        self.analyzed_sources = tuple(sources)
        labels = [source_label(source) for source in sources]
        self.reference_combo["values"] = labels
        if not labels:
            self.reference_file_var.set("")
            return
        if self.reference_file_var.get() not in labels:
            self.reference_file_var.set(labels[0])

    def fill_projection_table(self, infos: list[object]) -> None:
        for item in self.projection_tree.get_children():
            self.projection_tree.delete(item)
        for info in infos:
            self.projection_tree.insert("", END, values=(source_label(info.source), info.message))

    def choose_extra_reference(self) -> None:
        path = filedialog.askopenfilename(title="选择投影参考 shp", filetypes=[("Shapefile", "*.shp")])
        if path:
            self.reference_extra_var.set(path)
            self.reference_mode.set(1)
            self.update_target_projection()

    def update_target_projection(self) -> None:
        target_source = None
        if self.reference_mode.get() == 0:
            value = self.reference_file_var.get().strip()
            for source in self.analyzed_sources or tuple(self.selected_sources()):
                if source_label(source) == value:
                    target_source = source
                    break
        else:
            value = self.reference_extra_var.get().strip()
            if value:
                target_source = make_vector_source("shp", Path(value))
        self.target_source = target_source
        self.target_spatial_reference = None
        if hasattr(self, "projection_text"):
            self.projection_text.delete("1.0", END)
        if target_source is None:
            if hasattr(self, "projection_text"):
                self.projection_text.insert(END, "尚未选择投影基准。")
            return
        try:
            info = vector_tool.read_source_spatial_reference(target_source)
            self.target_spatial_reference = info.spatial_reference
            if hasattr(self, "projection_text"):
                self.projection_text.insert(END, f"投影来源：{source_label(target_source)}\n\n{vector_tool.projection_text(info.spatial_reference)}")
        except Exception as exc:
            if hasattr(self, "projection_text"):
                self.projection_text.insert(END, f"投影读取失败：{exc}")

    def report_key(self) -> tuple:
        return (
            tuple(source_label(source) for source in self.selected_sources()),
            self.area_var.get(),
            self.overlap_mode_var.get(),
            self.reference_mode.get(),
            self.reference_extra_var.get(),
            source_label(self.target_source) if self.target_source else "",
        )

    def validate_current_input(self) -> None:
        sources = self.selected_sources()
        if not sources:
            messagebox.showwarning("提示", "请先勾选需要计算的高标准农田面数据。")
            return
        if not self.area_var.get().strip():
            messagebox.showwarning("提示", "请先选择国标二级农业区。")
            return
        try:
            self.log_status("开始审查投影、字段匹配、字段类型、空值和类别值。")
            self.refresh_reference_choices(sources)
            projection_infos, projections_same = vector_tool.analyze_projection_state(sources)
            self.projection_infos = projection_infos
            self.projections_same = projections_same
            self.fill_projection_table(projection_infos)
            self.update_target_projection()
            if self.target_source is None or self.target_spatial_reference is None:
                messagebox.showwarning("提示", "请先选择有效的统一投影。")
                return
            rule_set = load_rule_set(self.rules_dir, self.area_var.get())
            report = validate_high_standard_sources(
                sources,
                rule_set,
                self.target_source,
                self.target_spatial_reference,
            )
        except Exception as exc:
            messagebox.showerror("审查失败", str(exc))
            return
        self.last_report = report
        self.last_report_key = self.report_key()
        self.show_report(report, ask_continue=False)
        if report.ok:
            self.log_status("审查通过，可以提交第一步计算任务。")
        else:
            self.log_status("审查未通过，已在报告中列出问题。")

    def show_report(self, report: HighStandardPreflightReport, ask_continue: bool) -> bool:
        window = Toplevel(self.root)
        window.title("第一步高标隶属度计算前审查报告")
        window.geometry("1040x720")
        text_frame = ttk.Frame(window)
        text_frame.pack(fill=BOTH, expand=True, padx=8, pady=8)
        text_frame.rowconfigure(0, weight=1)
        text_frame.columnconfigure(0, weight=1)
        text = Text(text_frame, wrap="word")
        text_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=text_scroll.set)
        text.grid(row=0, column=0, sticky="nsew")
        text_scroll.grid(row=0, column=1, sticky="ns")
        text.insert(END, report.text)
        text.configure(state="disabled")
        result = {"confirmed": False}

        button_row = ttk.Frame(window)
        button_row.pack(fill="x", padx=8, pady=8)

        def confirm() -> None:
            result["confirmed"] = True
            window.destroy()

        def cancel() -> None:
            result["confirmed"] = False
            window.destroy()

        if ask_continue:
            ttk.Button(button_row, text="确认无误，继续计算", command=confirm).pack(side=LEFT, padx=5)
            ttk.Button(button_row, text="取消", command=cancel).pack(side=LEFT, padx=5)
            window.transient(self.root)
            window.grab_set()
            self.root.wait_window(window)
            return result["confirmed"]
        ttk.Button(button_row, text="关闭", command=window.destroy).pack(side=LEFT, padx=5)
        return False

    def output_settings(self) -> tuple[str, Path, str | None] | None:
        gdb_text = self.output_gdb_var.get().strip()
        feature_name = self.output_feature_var.get().strip()
        if not gdb_text or not feature_name:
            messagebox.showwarning("提示", "请选择 GDB 并填写面要素类名。")
            return None
        output_path = Path(gdb_text if gdb_text.lower().endswith(".gdb") else f"{gdb_text}.gdb")
        output_feature_name = validate_gdb_feature_name(feature_name, output_path)
        if output_feature_name != feature_name:
            self.output_feature_var.set(output_feature_name)
            self.log_status(f"输出要素类名已按 FileGDB 规则修正为：{output_feature_name}")
        return "gdb", output_path, output_feature_name

    def submit_job(self) -> None:
        sources = self.selected_sources()
        if not sources:
            messagebox.showwarning("提示", "请先勾选需要计算的高标准农田面数据。")
            return
        if not self.area_var.get().strip():
            messagebox.showwarning("提示", "请先选择国标二级农业区。")
            return
        output = self.output_settings()
        if output is None:
            return
        output_kind, output_path, output_feature_name = output
        output_dataset = output_dataset_path(output_kind, output_path, output_feature_name)
        for source in sources:
            input_dataset = Path(source_dataset_path(source)).resolve()
            if output_dataset.resolve() == input_dataset:
                messagebox.showerror("输出错误", f"输出结果不能覆盖输入数据：{source_label(source)}")
                return
            if source.kind == "gdb" and output_path.resolve() == source.source_path.resolve():
                self.log_status(f"输出将保存到输入数据所在 GDB 的新图层：{output_feature_name}")
        try:
            report = self.last_report
            if report is None or self.last_report_key != self.report_key():
                self.log_status("当前输入没有最新审查报告，正在重新审查。")
                self.validate_current_input()
                report = self.last_report
            if report is None:
                return
            if not report.ok:
                self.show_report(report, ask_continue=False)
                messagebox.showerror("检查未通过", "输入数据存在问题，不能计算。请先按报告修正。")
                return
        except Exception as exc:
            messagebox.showerror("提交失败", str(exc))
            return
        if not self.show_report(report, ask_continue=True):
            self.log_status("用户取消计算任务，未提交。")
            return
        job = HighStandardCalculationJob(
            job_id=uuid.uuid4().hex[:8],
            input_sources=sources,
            output_path=output_path,
            output_feature_name=output_feature_name,
            output_kind=output_kind,
            overlap_mode=self.overlap_mode_var.get(),
            rule_set=report.rule_set,
            source_results=report.source_results,
            target_spatial_reference=report.target_spatial_reference,
            target_projection_source=report.target_projection_source,
            created_at=now_text(),
            validation_report=report.text,
        )
        self.job_queue.put(job)
        self.log_status(f"已提交任务 {job.job_id}，后台执行中。")

    def poll_worker_events(self) -> None:
        try:
            while True:
                event_type, payload = self.event_queue.get_nowait()
                self.log_status(f"[{payload.get('job_id')}] {payload.get('message')}")
                if event_type == "job_done" and self.on_job_done:
                    self.on_job_done(payload)
                if event_type in {"job_done", "job_failed"} and payload.get("log_path"):
                    self.log_status(f"日志：{payload['log_path']}")
        except queue.Empty:
            pass
        self.root.after(200, self.poll_worker_events)

    def log_status(self, message: str) -> None:
        self.status_text.insert(END, f"{now_text()}  {message}\n")
        self.status_text.see(END)

    def show_history(self) -> None:
        history_path = self.logs_dir / "membership_history.jsonl"
        if not history_path.exists():
            messagebox.showinfo("历史记录", "暂无历史记录。")
            return
        window = Toplevel(self.root)
        window.title("第一步高标隶属度计算历史记录")
        window.geometry("980x620")
        text_frame = ttk.Frame(window)
        text_frame.pack(fill=BOTH, expand=True, padx=8, pady=8)
        text_frame.rowconfigure(0, weight=1)
        text_frame.columnconfigure(0, weight=1)
        text = Text(text_frame, wrap="word")
        text_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=text_scroll.set)
        text.grid(row=0, column=0, sticky="nsew")
        text_scroll.grid(row=0, column=1, sticky="ns")
        for line in history_path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                text.insert(END, line + "\n")
                continue
            text.insert(
                END,
                f"任务 {record.get('job_id')} | {record.get('status')} | {record.get('created_at')}\n"
                f"输入：{record.get('input_sources')}\n"
                f"输出：{record.get('output_path')}\n"
                f"规则：{record.get('rule_path')}\n"
                f"重叠处理：{record.get('overlap_mode') or ''}\n"
                f"统计：{record.get('calculated_count')}\n"
                f"日志：{record.get('log_path')}\n"
                f"错误：{record.get('error') or ''}\n\n",
            )
        text.configure(state="disabled")

    def open_logs_folder(self) -> None:
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        try:
            import os

            os.startfile(self.logs_dir)
        except Exception as exc:
            messagebox.showerror("打开失败", str(exc))

    def delete_logs(self) -> None:
        if not messagebox.askyesno("确认删除", "确定删除第一步高标隶属度计算日志和历史记录吗？"):
            return
        deleted = 0
        for path in self.logs_dir.glob("membership_*.log"):
            path.unlink()
            deleted += 1
        history_path = self.logs_dir / "membership_history.jsonl"
        if history_path.exists():
            history_path.unlink()
            deleted += 1
        self.log_status(f"已删除 {deleted} 个第一步高标隶属度计算日志/历史文件。")

    def reset_inputs(self) -> None:
        self.area_options = self.load_area_options()
        if hasattr(self, "area_combo"):
            self.area_combo["values"] = self.area_options
        self.area_var.set(self.area_options[0] if self.area_options else DEFAULT_AREA_NAME)
        self.folder_var.set("")
        self.reference_mode.set(0)
        self.reference_file_var.set("")
        self.reference_extra_var.set("")
        self.output_gdb_var.set("")
        self.output_feature_var.set("Step1_高标隶属度")
        self.overlap_mode_var.set(OVERLAP_MODE_MARK)
        self.discovered_sources = []
        self.analyzed_sources = ()
        self.projections_same = False
        self.projection_infos = []
        self.target_spatial_reference = None
        self.target_source = None
        self.last_report = None
        self.last_report_key = None
        if hasattr(self, "source_checklist"):
            self.source_checklist.set_sources([])
        if hasattr(self, "reference_combo"):
            self.reference_combo["values"] = []
        if hasattr(self, "projection_tree"):
            for item in self.projection_tree.get_children():
                self.projection_tree.delete(item)
        if hasattr(self, "projection_text"):
            self.projection_text.delete("1.0", END)
            self.projection_text.insert(END, "选择并审查输入数据后，这里会显示统一投影。")
        self.log_status("第一步输入和参数已恢复为启动默认值。")


def main() -> int:
    try:
        require_runtime()
    except Exception as exc:
        root = Tk()
        root.withdraw()
        messagebox.showerror("缺少运行环境", str(exc))
        return 1
    root = Tk()
    MembershipApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
